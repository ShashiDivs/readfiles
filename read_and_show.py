"""
R2R Read and Show — Multi-Bank Extract + Match

1. Discovers how many banks are in the Excel files using GPT-4o
2. Extracts structured JSON per bank from bank statement and GL entries
3. Matches each bank's transactions — matched / mismatch / timing diff / missing
4. Logs every step with per-bank correlation ID
5. Prints consolidated summary across all banks

Usage:
  python read_and_show.py
"""
import json
import os
import sys
import time
import uuid
from pathlib import Path
from openai import AzureOpenAI
from dotenv import load_dotenv
import openpyxl
import logging
from datetime import datetime, timezone

load_dotenv()
client = AzureOpenAI(
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version="2024-02-01"
)
DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")

PROMPTS_DIR = Path(__file__).parent / "prompts"
DATA_DIR    = Path(__file__).parent / "data"
LOGS_DIR    = Path(__file__).parent / "logs"

CHUNK_SIZE  = 150  # smaller chunks keep output within max_tokens limit


# ── Logger ─────────────────────────────────────────────────────────────────────

class JSONFormatter(logging.Formatter):
    def __init__(self, correlation_id: str):
        super().__init__()
        self.correlation_id = correlation_id

    def format(self, record: logging.LogRecord) -> str:
        entry = {
            "timestamp":      datetime.now(timezone.utc).isoformat(),
            "level":          record.levelname,
            "correlation_id": self.correlation_id,
            "station":        getattr(record, "station", "pipeline"),
            "message":        record.getMessage()
        }
        for key, val in record.__dict__.items():
            if key not in ("msg", "args", "levelname", "levelno", "pathname",
                           "filename", "module", "exc_info", "exc_text",
                           "stack_info", "lineno", "funcName", "created",
                           "msecs", "relativeCreated", "thread", "threadName",
                           "processName", "process", "name", "message", "station"):
                entry[key] = val
        return json.dumps(entry)


def setup_logger(correlation_id: str) -> logging.Logger:
    LOGS_DIR.mkdir(exist_ok=True)
    log_file = LOGS_DIR / f"{correlation_id}.log"

    logger = logging.getLogger(f"r2r.{correlation_id}")
    logger.handlers.clear()

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(JSONFormatter(correlation_id))
    logger.addHandler(console_handler)

    file_handler = logging.FileHandler(log_file)
    file_handler.setFormatter(JSONFormatter(correlation_id))
    logger.addHandler(file_handler)

    logger.setLevel(logging.DEBUG)
    logger.info(f"Log file: {log_file}")
    return logger


def get_log(station: str, logger: logging.Logger) -> logging.LoggerAdapter:
    return logging.LoggerAdapter(logger, extra={"station": station})


# ── Helpers ────────────────────────────────────────────────────────────────────

def load_prompt(file_type: str) -> str:
    prompt_file = PROMPTS_DIR / f"{file_type}.txt"
    if not prompt_file.exists():
        print(f"Prompt file not found: {prompt_file}")
        sys.exit(1)
    return prompt_file.read_text().strip()


def read_excel_rows(file_path: str, log, sheet_name: str = None) -> tuple:
    """
    Read Excel rows from a specific sheet or all sheets.
    Returns (header_text, all_rows).
    """
    log.info("Reading Excel file", extra={"file": Path(file_path).name, "sheet": sheet_name or "ALL"})
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_rows = []

    sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames

    for sheet in sheets:
        ws = wb[sheet]
        all_rows.append(f"=== Sheet: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                all_rows.append(" | ".join(str(c) if c is not None else "" for c in row))

    header_text = "\n".join(all_rows[:20])
    log.info(
        "Excel read complete",
        extra={"file": Path(file_path).name, "total_rows": len(all_rows), "sheet": sheet_name or "ALL"}
    )
    return header_text, all_rows


def read_excel_summary(file_path: str) -> str:
    """Read first 50 rows from all sheets — used for bank discovery."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"=== Sheet: {sheet} ===")
        count = 0
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                count += 1
                if count >= 50:
                    break
    return "\n".join(lines)


# ── Discovery ──────────────────────────────────────────────────────────────────

def discover_banks(file_path: str, log) -> list:
    """
    Ask GPT-4o to identify how many banks are in the file and where each one is.
    Returns list of dicts: [{"name": "HDFC", "sheet": "HDFC"}, ...]
    """
    log.info("Discovering banks in file", extra={"file": Path(file_path).name})
    start = time.time()

    summary = read_excel_summary(file_path)

    response = client.chat.completions.create(
        model=DEPLOYMENT,
        messages=[
            {"role": "system", "content": load_prompt("discover")},
            {"role": "user",   "content": f"Excel file content (first 50 rows per sheet):\n\n{summary}\n\nReturn valid JSON only."}
        ],
        temperature=0,
        response_format={"type": "json_object"}
    )

    result = json.loads(response.choices[0].message.content)
    banks  = result.get("banks", [])
    duration_ms = round((time.time() - start) * 1000)

    log.info(
        "Discovery complete",
        extra={
            "file":        Path(file_path).name,
            "total_banks": result.get("total_banks", len(banks)),
            "structure":   result.get("structure"),
            "banks":       [b["name"] for b in banks],
            "duration_ms": duration_ms
        }
    )
    return banks


# ── Extraction ─────────────────────────────────────────────────────────────────

def extract_chunk(chunk_text: str, file_type: str, chunk_no: int, total_chunks: int, log) -> dict:
    log.info(
        "Extracting chunk",
        extra={"file_type": file_type, "chunk": f"{chunk_no}/{total_chunks}"}
    )
    start = time.time()

    response = client.chat.completions.create(
        model=DEPLOYMENT,
        messages=[
            {"role": "system", "content": load_prompt(file_type)},
            {"role": "user",   "content": f"Raw Excel data (chunk {chunk_no} of {total_chunks}):\n\n{chunk_text}\n\nReturn valid JSON only."}
        ],
        temperature=0,
        max_tokens=8192,
        response_format={"type": "json_object"}
    )

    duration_ms = round((time.time() - start) * 1000)
    usage = response.usage
    finish_reason = response.choices[0].finish_reason

    log.info(
        "Chunk extraction complete",
        extra={
            "chunk":         f"{chunk_no}/{total_chunks}",
            "input_tokens":  usage.prompt_tokens,
            "output_tokens": usage.completion_tokens,
            "finish_reason": finish_reason,
            "duration_ms":   duration_ms
        }
    )

    if finish_reason == "length":
        log.warning(
            "Response truncated — output hit max_tokens limit, reduce CHUNK_SIZE if this persists",
            extra={"chunk": f"{chunk_no}/{total_chunks}"}
        )

    try:
        return json.loads(response.choices[0].message.content)
    except json.JSONDecodeError as e:
        log.error(
            "JSON parse failed — response likely truncated",
            extra={"chunk": f"{chunk_no}/{total_chunks}", "error": str(e)}
        )
        # Return empty structure so pipeline continues with other chunks
        tx_key = "transactions" if file_type == "bank_statement" else "entries"
        return {tx_key: []}


def merge_results(results: list, file_type: str) -> dict:
    if not results:
        return {}
    merged = results[0].copy()
    tx_key = "transactions" if file_type == "bank_statement" else "entries"
    all_tx = []
    for result in results:
        all_tx.extend(result.get(tx_key, []))
    merged[tx_key] = all_tx
    return merged


def extract(file_path: str, file_type: str, log, sheet_name: str = None) -> dict:
    """Extract structured JSON from Excel. Handles chunking for large files."""
    header_text, all_rows = read_excel_rows(file_path, log, sheet_name=sheet_name)

    if len(all_rows) <= CHUNK_SIZE:
        log.info("File fits in single call", extra={"rows": len(all_rows)})
        return extract_chunk("\n".join(all_rows), file_type, 1, 1, log)

    data_rows = all_rows[20:]
    chunks = []
    for i in range(0, len(data_rows), CHUNK_SIZE):
        chunk = header_text + "\n" + "\n".join(data_rows[i:i + CHUNK_SIZE])
        chunks.append(chunk)

    total_chunks = len(chunks)
    log.info(
        "Chunking file",
        extra={"total_rows": len(all_rows), "chunks": total_chunks, "chunk_size": CHUNK_SIZE}
    )

    results = [extract_chunk(chunk, file_type, i + 1, total_chunks, log) for i, chunk in enumerate(chunks)]
    merged  = merge_results(results, file_type)

    tx_key = "transactions" if file_type == "bank_statement" else "entries"
    log.info(
        "Extraction complete",
        extra={"chunks_processed": total_chunks, "total_transactions": len(merged.get(tx_key, []))}
    )
    return merged


# ── Matching ───────────────────────────────────────────────────────────────────

def match(bank_data: dict, gl_data: dict, log) -> dict:
    log.info("Starting matching — bank statement vs GL entries")
    start = time.time()

    prompt = f"""You are a bank reconciliation expert.

Compare these two extracted datasets:
1. Bank Statement — transactions recorded by the bank
2. GL Entries — transactions recorded in our accounting system

Match each transaction by reference number, amount and narration.
Use fuzzy matching for narrations — the same transaction will look different in each system.
For example: "NEFT CR INFOSYS LIMITED" matches "Vendor payment — Infosys Ltd"

For every transaction classify it as one of:
- MATCHED          : same transaction, same amount on both sides
- MISMATCH         : same transaction but different amount — note the difference
- TIMING_DIFFERENCE: exists on one side only but expected (cheque not cleared, payment in transit)
- MISSING          : exists on one side only with no clear explanation — needs investigation

Return JSON only:
{{
  "matched": [
    {{
      "ref": "transaction reference",
      "description": "what this transaction is",
      "gl_amount": 0.00,
      "bank_amount": 0.00
    }}
  ],
  "mismatches": [
    {{
      "ref": "...",
      "description": "...",
      "gl_amount": 0.00,
      "bank_amount": 0.00,
      "difference": 0.00,
      "likely_reason": "why this difference might exist"
    }}
  ],
  "timing_differences": [
    {{
      "ref": "...",
      "description": "...",
      "amount": 0.00,
      "present_in": "GL_ONLY or BANK_ONLY",
      "reason": "why this is likely a timing difference"
    }}
  ],
  "missing": [
    {{
      "ref": "...",
      "description": "...",
      "amount": 0.00,
      "present_in": "GL_ONLY or BANK_ONLY",
      "action": "what finance team should do"
    }}
  ],
  "summary": {{
    "total_matched": 0,
    "total_mismatches": 0,
    "total_timing_differences": 0,
    "total_missing": 0,
    "is_reconciled": true,
    "commentary": "one line plain English summary of the reconciliation result"
  }}
}}

Bank Statement:
{json.dumps(bank_data, indent=2)}

GL Entries:
{json.dumps(gl_data, indent=2)}"""

    response = client.chat.completions.create(
        model=DEPLOYMENT,
        messages=[
            {"role": "system", "content": "You are a bank reconciliation expert. Return valid JSON only."},
            {"role": "user",   "content": prompt}
        ],
        temperature=0,
        response_format={"type": "json_object"}
    )

    duration_ms = round((time.time() - start) * 1000)
    usage   = response.usage
    result  = json.loads(response.choices[0].message.content)
    summary = result.get("summary", {})

    log.info(
        "Matching complete",
        extra={
            "matched":            summary.get("total_matched", 0),
            "mismatches":         summary.get("total_mismatches", 0),
            "timing_differences": summary.get("total_timing_differences", 0),
            "missing":            summary.get("total_missing", 0),
            "is_reconciled":      summary.get("is_reconciled"),
            "input_tokens":       usage.prompt_tokens,
            "output_tokens":      usage.completion_tokens,
            "duration_ms":        duration_ms
        }
    )

    for m in result.get("mismatches", []):
        log.warning("Mismatch found", extra={"ref": m.get("ref"), "difference": m.get("difference"), "reason": m.get("likely_reason")})

    for m in result.get("missing", []):
        log.warning("Missing entry", extra={"ref": m.get("ref"), "present_in": m.get("present_in"), "action": m.get("action")})

    for t in result.get("timing_differences", []):
        log.info("Timing difference noted", extra={"ref": t.get("ref"), "present_in": t.get("present_in"), "reason": t.get("reason")})

    return result


# ── Print helpers ──────────────────────────────────────────────────────────────

def print_extraction(label: str, data: dict):
    print(f"\n{'─'*60}")
    print(f"EXTRACTED — {label}")
    print(f"{'─'*60}")
    print(json.dumps(data, indent=2))


def print_match_result(bank_name: str, result: dict):
    summary = result.get("summary", {})
    reconciled = summary.get("is_reconciled")
    status = "RECONCILED" if reconciled else "NOT RECONCILED"

    print(f"\n{'─'*60}")
    print(f"MATCH RESULT — {bank_name}  [{status}]")
    print(f"{'─'*60}")
    print(f"  Matched            : {summary.get('total_matched', 0)}")
    print(f"  Mismatches         : {summary.get('total_mismatches', 0)}")
    print(f"  Timing differences : {summary.get('total_timing_differences', 0)}")
    print(f"  Missing entries    : {summary.get('total_missing', 0)}")
    print(f"  Commentary         : {summary.get('commentary')}")

    if result.get("matched"):
        print(f"\n  MATCHED ({len(result['matched'])}):")
        for m in result["matched"]:
            print(f"    ✓  {m.get('ref',''):<12} {m.get('description','')[:45]:<45}  {m.get('gl_amount', 0):>12,.0f}")

    if result.get("mismatches"):
        print(f"\n  MISMATCHES ({len(result['mismatches'])}):")
        for m in result["mismatches"]:
            print(f"    ✗  {m.get('ref',''):<12} {m.get('description','')[:35]:<35}  GL:{m.get('gl_amount',0):>10,.0f}  Bank:{m.get('bank_amount',0):>10,.0f}  Diff:{m.get('difference',0):>8,.0f}")
            print(f"       Reason: {m.get('likely_reason','')}")

    if result.get("timing_differences"):
        print(f"\n  TIMING DIFFERENCES ({len(result['timing_differences'])}):")
        for t in result["timing_differences"]:
            print(f"    ⏳  {t.get('ref',''):<12} {t.get('description','')[:40]:<40}  {t.get('amount',0):>12,.0f}  ({t.get('present_in','')})")
            print(f"        Reason: {t.get('reason','')}")

    if result.get("missing"):
        print(f"\n  MISSING ENTRIES ({len(result['missing'])}):")
        for m in result["missing"]:
            print(f"    ?  {m.get('ref',''):<12} {m.get('description','')[:40]:<40}  {m.get('amount',0):>12,.0f}  ({m.get('present_in','')})")
            print(f"       Action: {m.get('action','')}")


def print_consolidated_summary(results: list):
    print(f"\n{'='*60}")
    print("CONSOLIDATED SUMMARY — ALL BANKS")
    print(f"{'='*60}")
    print(f"  {'Bank':<15} {'Matched':>8} {'Mismatch':>9} {'Timing':>7} {'Missing':>8}  Status")
    print(f"  {'─'*15} {'─'*8} {'─'*9} {'─'*7} {'─'*8}  {'─'*14}")
    all_reconciled = True
    for r in results:
        s = r["match_result"].get("summary", {})
        reconciled = s.get("is_reconciled", False)
        if not reconciled:
            all_reconciled = False
        status = "✓ RECONCILED" if reconciled else "✗ OPEN ITEMS"
        print(f"  {r['bank']:<15} {s.get('total_matched',0):>8} {s.get('total_mismatches',0):>9} {s.get('total_timing_differences',0):>7} {s.get('total_missing',0):>8}  {status}")
    print(f"{'─'*60}")
    overall = "ALL BANKS RECONCILED" if all_reconciled else "PERIOD CLOSE BLOCKED — OPEN ITEMS EXIST"
    print(f"  Overall: {overall}")
    print(f"{'='*60}\n")


# ── Main ───────────────────────────────────────────────────────────────────────

def run():
    session_id = f"R2R-DEMO-{uuid.uuid4().hex[:8]}"

    print(f"\n{'='*60}")
    print("R2R — Multi-Bank Extract and Match")
    print(f"{'='*60}")
    print(f"  Session ID : {session_id}")

    # Session-level logger for discovery
    session_logger = setup_logger(session_id)
    log_main = get_log("pipeline", session_logger)
    log_main.info("Session started", extra={"session_id": session_id})

    start_total = time.time()

    # ── Check files exist ──────────────────────────────────────────────────────
    bank_file = DATA_DIR / "bank_statement.xlsx"
    gl_file   = DATA_DIR / "bank_gl_entries.xlsx"

    for f in [bank_file, gl_file]:
        if not f.exists():
            log_main.error("File not found", extra={"path": str(f)})
            print(f"\n  Missing file: {f}")
            sys.exit(1)

    # ── Discover banks in both files ───────────────────────────────────────────
    print(f"\n  Discovering banks in statement file...")
    bank_banks = discover_banks(str(bank_file), get_log("discovery", session_logger))

    print(f"  Discovering banks in GL file...")
    gl_banks   = discover_banks(str(gl_file),   get_log("discovery", session_logger))

    # Build lookup: bank name → sheet name
    gl_lookup = {b["name"].upper(): b.get("sheet") for b in gl_banks}

    print(f"\n  Banks found in statement : {[b['name'] for b in bank_banks]}")
    print(f"  Banks found in GL        : {list(gl_lookup.keys())}")

    all_results = []

    # ── Process each bank ──────────────────────────────────────────────────────
    for bank_info in bank_banks:
        bank_name  = bank_info["name"]
        bank_sheet = bank_info.get("sheet")

        # Pairing strategy:
        # 1. Exact name match (case-insensitive)
        # 2. Partial name match — bank name contains GL name or vice versa
        # 3. If only one bank on each side, pair them directly regardless of name
        gl_sheet = gl_lookup.get(bank_name.upper())

        if not gl_sheet:
            # Try partial match — "JPMorgan Chase Bank, NA Mumbai" contains "JPMORGAN"
            for gl_name, sheet in gl_lookup.items():
                if gl_name in bank_name.upper() or bank_name.upper() in gl_name:
                    gl_sheet = sheet
                    log_main.info(
                        "Paired by partial name match",
                        extra={"bank": bank_name, "gl_name": gl_name}
                    )
                    break

        if not gl_sheet and len(bank_banks) == 1 and len(gl_banks) == 1:
            # Single bank on each side — pair them directly
            gl_sheet = gl_banks[0].get("sheet")
            log_main.info(
                "Paired by single-bank fallback",
                extra={"bank": bank_name, "gl_name": gl_banks[0]["name"]}
            )

        if not gl_sheet:
            log_main.warning(
                "No matching GL entries found for bank — skipping",
                extra={"bank": bank_name, "available_gl_banks": list(gl_lookup.keys())}
            )
            print(f"\n  ⚠  No GL entries found for {bank_name} — skipping")
            continue

        # Per-bank correlation ID and logger
        correlation_id = f"R2R-{bank_name.upper()}-{uuid.uuid4().hex[:8]}"
        bank_logger    = setup_logger(correlation_id)
        log_extract    = get_log("extraction", bank_logger)
        log_match      = get_log("matching",   bank_logger)

        print(f"\n{'─'*60}")
        print(f"  Processing bank: {bank_name}  [{correlation_id}]")
        print(f"{'─'*60}")

        # Extract
        bank_data = extract(str(bank_file), "bank_statement", log_extract, sheet_name=bank_sheet)
        bank_tx_count = len(bank_data.get("transactions", []))
        print(f"  Extracted {bank_tx_count} transactions from bank statement")
        print_extraction(f"BANK STATEMENT — {bank_name}", bank_data)

        gl_data = extract(str(gl_file), "bank_gl", log_extract, sheet_name=gl_sheet)
        gl_tx_count = len(gl_data.get("entries", []))
        print(f"  Extracted {gl_tx_count} entries from GL")
        print_extraction(f"GL ENTRIES — {bank_name}", gl_data)

        # Match
        match_result = match(bank_data, gl_data, log_match)
        print_match_result(bank_name, match_result)

        all_results.append({
            "bank":           bank_name,
            "correlation_id": correlation_id,
            "bank_statement": bank_data,
            "gl_entries":     gl_data,
            "match_result":   match_result
        })

    # ── Consolidated summary ───────────────────────────────────────────────────
    if all_results:
        print_consolidated_summary(all_results)

    # ── Save output ────────────────────────────────────────────────────────────
    duration_ms = round((time.time() - start_total) * 1000)

    output = {
        "session_id":  session_id,
        "total_banks": len(all_results),
        "banks":       all_results
    }

    out_path = DATA_DIR / "output.json"
    out_path.write_text(json.dumps(output, indent=2))

    log_main.info(
        "Session complete",
        extra={"duration_ms": duration_ms, "banks_processed": len(all_results), "output": str(out_path)}
    )

    print(f"  Duration : {duration_ms}ms")
    print(f"  Output   : {out_path}")
    print(f"  Logs     : {LOGS_DIR}/\n")


if __name__ == "__main__":
    run()
